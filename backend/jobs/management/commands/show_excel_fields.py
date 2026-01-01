"""
Management command to show Excel file structure
Usage: python manage.py show_excel_fields <excel_file_path>
"""
from django.core.management.base import BaseCommand, CommandError
import openpyxl
import os


class Command(BaseCommand):
    help = 'Show Excel file structure and fields'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Path to Excel file')

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        
        if not os.path.exists(excel_file):
            raise CommandError(f'Excel file not found: {excel_file}')

        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            self.stdout.write(self.style.SUCCESS('=' * 60))
            self.stdout.write(self.style.SUCCESS('Excel File Structure'))
            self.stdout.write(self.style.SUCCESS('=' * 60))
            self.stdout.write(f'\nSheet: {sheet.title}')
            self.stdout.write(f'Total Rows: {sheet.max_row}')
            self.stdout.write(f'Total Columns: {sheet.max_column}')
            
            # Get headers
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else "")
            
            self.stdout.write(f'\nColumn Headers ({len(headers)} columns):')
            for idx, header in enumerate(headers, start=1):
                self.stdout.write(f'  {idx}. {header}')
            
            # Show first few data rows
            self.stdout.write(f'\nFirst 3 Data Rows:')
            for row_num in range(2, min(5, sheet.max_row + 1)):
                row = sheet[row_num]
                self.stdout.write(f'\nRow {row_num}:')
                for idx, cell in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        value = str(cell.value) if cell.value else ""
                        # Truncate long values
                        if len(value) > 100:
                            value = value[:100] + "..."
                        self.stdout.write(f'  {headers[idx]}: {value}')
            
            self.stdout.write(self.style.SUCCESS('\n' + '=' * 60))
            
        except ImportError:
            raise CommandError('openpyxl is not installed. Run: pip install openpyxl')
        except Exception as e:
            raise CommandError(f'Error reading Excel file: {str(e)}')

