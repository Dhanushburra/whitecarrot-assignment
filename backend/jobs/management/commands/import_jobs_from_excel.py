"""
Management command to import jobs from Excel file
Usage: python manage.py import_jobs_from_excel <company_slug> <excel_file_path> [--limit N]
"""
from django.core.management.base import BaseCommand, CommandError
from companies.models import Company
from jobs.models import Job
import openpyxl
import os


class Command(BaseCommand):
    help = 'Import jobs from Excel file to a specific company'

    def add_arguments(self, parser):
        parser.add_argument('company_slug', type=str, help='Company slug (e.g., accenture)')
        parser.add_argument('excel_file', type=str, help='Path to Excel file')
        parser.add_argument(
            '--limit',
            type=int,
            default=None,
            help='Limit number of jobs to import (e.g., --limit 10)',
        )
        parser.add_argument(
            '--show-fields',
            action='store_true',
            help='Show Excel file structure and exit',
        )

    def handle(self, *args, **options):
        company_slug = options['company_slug']
        excel_file = options['excel_file']
        limit = options['limit']
        show_fields = options['show_fields']

        # Check if file exists
        if not os.path.exists(excel_file):
            raise CommandError(f'Excel file not found: {excel_file}')

        # Read Excel file
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            # Get headers
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            
            # Show fields if requested
            if show_fields:
                self.stdout.write(self.style.SUCCESS('=' * 60))
                self.stdout.write(self.style.SUCCESS('Excel File Structure'))
                self.stdout.write(self.style.SUCCESS('=' * 60))
                self.stdout.write(f'\nSheet: {sheet.title}')
                self.stdout.write(f'Total Rows: {sheet.max_row}')
                self.stdout.write(f'Total Columns: {sheet.max_column}')
                self.stdout.write(f'\nColumn Headers:')
                for idx, header in enumerate(headers, start=1):
                    self.stdout.write(f'  {idx}. {header}')
                self.stdout.write(f'\nFirst Data Row Sample:')
                if sheet.max_row > 1:
                    row = sheet[2]
                    for idx, cell in enumerate(row):
                        if idx < len(headers) and headers[idx]:
                            value = str(cell.value) if cell.value else ""
                            if len(value) > 50:
                                value = value[:50] + "..."
                            self.stdout.write(f'  {headers[idx]}: {value}')
                return
            
            # Get company
            try:
                company = Company.objects.get(slug=company_slug)
                self.stdout.write(self.style.SUCCESS(f'Found company: {company.name}'))
            except Company.DoesNotExist:
                raise CommandError(f'Company with slug "{company_slug}" not found')

            # Map Excel columns to job fields (flexible mapping)
            column_map = {}
            for idx, header in enumerate(headers, start=1):
                header_lower = str(header).lower().strip()
                if not header_lower:
                    continue
                    
                # Flexible column matching
                if any(x in header_lower for x in ['title', 'job title', 'position', 'role']):
                    column_map['title'] = idx
                elif any(x in header_lower for x in ['description', 'job description', 'details']):
                    column_map['description'] = idx
                elif any(x in header_lower for x in ['location', 'city', 'address']):
                    column_map['location'] = idx
                elif any(x in header_lower for x in ['work policy', 'work_policy', 'policy', 'remote', 'hybrid', 'onsite']):
                    column_map['work_policy'] = idx
                elif any(x in header_lower for x in ['employment', 'employment type', 'employement', 'employement type']):
                    column_map['employment_type'] = idx
                elif any(x in header_lower for x in ['type', 'job type']):
                    column_map['employment_type'] = idx  # Fallback
                elif any(x in header_lower for x in ['department', 'dept', 'team', 'division']):
                    column_map['department'] = idx
                elif any(x in header_lower for x in ['experience', 'level', 'senior', 'junior', 'mid']):
                    column_map['experience'] = idx
                elif any(x in header_lower for x in ['salary', 'compensation', 'pay', 'wage', 'salary range']):
                    column_map['salary_range'] = idx
                elif any(x in header_lower for x in ['posted', 'date', 'posted date', 'posted_date']):
                    column_map['posted_date'] = idx
            
            self.stdout.write(f'\nColumn mapping:')
            for field, col_idx in column_map.items():
                self.stdout.write(f'  {field}: Column {col_idx} ({headers[col_idx-1]})')
            
            # Check required fields
            if 'title' not in column_map:
                raise CommandError('Could not find "title" or "job title" column in Excel file')
            
            # Read data rows
            imported_count = 0
            skipped_count = 0
            total_rows = sheet.max_row - 1  # Exclude header
            
            # Determine how many rows to process
            rows_to_process = limit if limit else total_rows
            rows_to_process = min(rows_to_process, total_rows)
            
            self.stdout.write(f'\nProcessing {rows_to_process} rows...\n')
            
            for row_num in range(2, min(2 + rows_to_process, sheet.max_row + 1)):
                row = sheet[row_num]
                
                # Extract data
                job_data = {
                    'company': company,
                    'title': self._get_cell_value(row, column_map.get('title')),
                    'location': self._get_cell_value(row, column_map.get('location')) or 'Not specified',
                    'work_policy': self._normalize_work_policy(self._get_cell_value(row, column_map.get('work_policy'))),
                    'employment_type': self._normalize_employment_type(self._get_cell_value(row, column_map.get('employment_type'))),
                }
                
                # Optional fields
                description = self._get_cell_value(row, column_map.get('description'))
                if description and description.strip():
                    job_data['description'] = description.strip()
                
                department = self._get_cell_value(row, column_map.get('department'))
                if department and department.strip():
                    job_data['department'] = department.strip()
                
                experience = self._normalize_experience(self._get_cell_value(row, column_map.get('experience')))
                if experience:
                    job_data['experience'] = experience
                
                salary_range = self._get_cell_value(row, column_map.get('salary_range'))
                if salary_range and salary_range.strip():
                    job_data['salary_range'] = salary_range.strip()
                
                posted_date = self._get_cell_value(row, column_map.get('posted_date'))
                if posted_date and posted_date.strip():
                    job_data['posted_date'] = posted_date.strip()
                else:
                    job_data['posted_date'] = 'Just now'
                
                # Note: is_active field has been removed
                
                # Skip if title is missing
                if not job_data['title']:
                    skipped_count += 1
                    continue
                
                # Create job
                job, created = Job.objects.get_or_create(
                    company=company,
                    title=job_data['title'],
                    defaults=job_data
                )
                
                if created:
                    imported_count += 1
                    self.stdout.write(self.style.SUCCESS(f'  ✓ Imported: {job.title}'))
                else:
                    skipped_count += 1
                    self.stdout.write(self.style.WARNING(f'  - Skipped (already exists): {job.title}'))
            
            self.stdout.write(self.style.SUCCESS(
                f'\n✅ Import complete!'
                f'\n   Imported: {imported_count} jobs'
                f'\n   Skipped: {skipped_count} jobs'
            ))
            
        except ImportError:
            raise CommandError('openpyxl is not installed. Run: pip install openpyxl')
        except Exception as e:
            raise CommandError(f'Error: {str(e)}')

    def _get_cell_value(self, row, column_index):
        """Get cell value safely"""
        if column_index and column_index <= len(row):
            value = row[column_index - 1].value
            return str(value).strip() if value else None
        return None

    def _normalize_employment_type(self, employment_type):
        """Normalize employment type to match model choices"""
        if not employment_type:
            return 'full-time'
        
        emp_type_lower = str(employment_type).lower().strip()
        
        if any(x in emp_type_lower for x in ['full', 'full-time', 'fulltime']):
            return 'full-time'
        elif any(x in emp_type_lower for x in ['part', 'part-time', 'parttime']):
            return 'part-time'
        elif 'contract' in emp_type_lower:
            return 'contract'
        else:
            return 'full-time'  # Default
    
    def _normalize_work_policy(self, work_policy):
        """Normalize work policy to match model choices"""
        if not work_policy:
            return 'onsite'
        
        policy_lower = str(work_policy).lower().strip()
        
        if 'remote' in policy_lower:
            return 'remote'
        elif 'hybrid' in policy_lower:
            return 'hybrid'
        elif any(x in policy_lower for x in ['onsite', 'on-site', 'on site', 'office']):
            return 'onsite'
        else:
            return 'onsite'  # Default
    
    def _normalize_experience(self, experience):
        """Normalize experience level to match model choices"""
        if not experience:
            return None
        
        exp_lower = str(experience).lower().strip()
        
        # Check for exact matches first
        if exp_lower in ['senior', 'sr', 'lead', 'principal']:
            return 'senior'
        elif exp_lower in ['junior', 'jr', 'entry', 'associate']:
            return 'junior'
        elif exp_lower in ['mid-level', 'mid level', 'midlevel', 'mid', 'middle']:
            return 'mid-level'
        else:
            # Try partial matching
            if any(x in exp_lower for x in ['senior', 'sr', 'lead', 'principal']):
                return 'senior'
            elif any(x in exp_lower for x in ['junior', 'jr', 'entry', 'associate']):
                return 'junior'
            elif any(x in exp_lower for x in ['mid', 'middle']):
                return 'mid-level'
            else:
                return None  # Optional field

